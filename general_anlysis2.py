import PDF_save
from paddleocr import PaddleOCR
import re
import torch        #是下载torch而不是pytorch
from transformers import BertTokenizer, BertForSequenceClassification, AdamW, BertConfig
from torch.optim.lr_scheduler import StepLR
import multiprocessing
from datetime import datetime
import os
import re
from openpyxl import load_workbook
os.environ['KMP_DUPLICATE_LIB_OK']='True'


##全局变量
eyes_pattern = ["视力","色觉","辨色力"]
ear_pattern = ["外耳道","鼓膜","听力","耳"]
nose_pattern = ["鼻腔","鼻中隔","鼻"]
Xray_pattern = ["DR","胸部正位","全数字X光（DR）","胸部正位检查","胸片","胸部CT","胸廓两侧对称"]
ECG_pattern = ["心电图","静态心电图检查","ECG"]
ALT_pattern = ["谷丙转氨酶","ALT","血清丙氨酸氨基转移酶"]
AST_pattern = ["谷草转氨酶","AST","天冬氨酸氨基转移酶"]
GGT_pattern = ["谷氨酰转移酶","GGT","GCT","rGT"]
internal_pattern = ["内科"]
surgery_pattern = ["外科"]
BP_pattern = ["血压","BP","收缩压","舒张压"]
mouth_pattern = ["扁桃体","咽部","口咽"]
summary = ["医生建议","体检结论","体检报告总述","终检结论","结论及建议"]        #总结
conclusion = ["小结","结论"]                          #小结
TXT_location=[]

info = {"视力":[0,-1,-1],"听力":[0,-1,-1],"鼻":[0,-1,-1],"胸透":[0,-1,-1],"心电图":[0,-1,-1],
        "谷丙转氨酶":[0,-1,-1],"谷草转氨酶":[0,-1,-1],"谷氨酰转移酶":[0,-1,-1],"内科":[0,-1,-1],"外科":[0,-1,-1],
        "血压":[0,-1,-1],"咽部":[0,-1,-1],"医生总结":[0,-1,-1]}          #体检分段信息，第一个表示有没有做，后面两个表示始末位置

#pdf转图片
def extract_text_allpage(pdf_path,img_folder):
    imgs_path = PDF_save.pdf_to_images(pdf_path,img_folder)
    print("pdf转换成功!!,存储路径为{}".format(imgs_path))
    return imgs_path

#预测标签
def predict_label(sentence, model, tokenizer):
    # 对输入句子进行编码
    encoding = tokenizer.encode_plus(
        sentence,
        add_special_tokens=True,
        padding='max_length',
        truncation=True,
        max_length=128,
        return_tensors='pt'
    )

    input_ids = encoding['input_ids'].squeeze()
    attention_mask = encoding['attention_mask'].squeeze()

    # 在模型中进行前向传播
    with torch.no_grad():
        outputs = model(input_ids.unsqueeze(0), attention_mask=attention_mask.unsqueeze(0))
        logits = outputs.logits

    # 获取预测的标签
    _, predicted_label = torch.max(logits, dim=1)

    return predicted_label.item()

#提取日期并转换成对应格式
def extract_date(text):
    patterns = [
        r'(\d{4})年(\d{1,2})月(\d{1,2})日',      #2012年12月12日
        r'(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})', #2012-12-12
        r'(\d{4})(\d{1,2})(\d{1,2})'            #202110412
    ]

    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            groups = match.groups()
            # 根据不同的日期格式尝试转换为datetime对象
            try:
                year, month, day = map(int, groups)
                date_obj = datetime(year, month, day)
                return date_obj
            except ValueError:
                return None  # 若无法转换，则忽略并尝试下一个格式

# ocr识别   
def ocr(img_path):
    content = []
    location = []
    sorce = 0
    ocr = PaddleOCR(use_angle_cls=True, lang="ch")  # need to run only once to download and load model into memory
    result = ocr.ocr(img_path, cls=True)
    for idx in range(len(result)):
        res = result[idx]
        for line in res:
            if isinstance(line, list):
                location.append(line)
            else:
                content.append(line[0])
                if line[1]<0.85:
                    sorce+=1
    return result,content,location,sorce

#图片转文字
def img2word(img_paths):
    if type(img_paths) == str:               #格式为jpg的
        print(f"不迭代的情况下：现在传入格式为:{type(img_paths)}\n具体内容为:{img_paths}")
        content,location,sorce = ocr(img_paths)
    else:                                    #格式为pdf的
        result = []
        content = []
        location = []
        sorce = 0
        print(f"迭代的情况下：现在传入格式为:{type(img_paths)}\n具体内容为:{img_paths}")
        for img_path in img_paths:
            r,c,l,s = ocr(img_path)
            content += c
            location += l
            sorce += s
            result += r
    if (len(content)) == 0:
        print("未识别出内容,建议转人工查看！！！")
        content = ["1"]                     #这个占位置的，不用管
    return result,content,location,sorce/(len(content))

#检测项目是否做全【先利用关键字和出现的下标将区域分块，分块之后再利用关键词找小结或者结论】
def is_enough(sentence):
    index_list,doctor_index_list = []
    for j in range(len(sentence)):
        if bool([i for i in eyes_pattern  if i in sentence[j]]) and info["视力"][0]==0:     #找到第一次找到关键词
            info["视力"][0],info["视力"][1]=1,j     #更新项目状态，项目起始位置          
            index_list.append(j)
        elif bool([i for i in ear_pattern  if i in sentence[j]]) and info["听力"][0]==0:     #找到第一次找到关键词
            info["听力"][0],info["听力"][1]=1,j               
            index_list.append(j)
        elif bool([i for i in nose_pattern  if i in sentence[j]]) and info["鼻"][0]==0:     #找到第一次找到关键词
            info["鼻"][0],info["鼻"][1]=1,j               
            index_list.append(j)
        elif bool([i for i in Xray_pattern  if i in sentence[j]]) and info["胸透"][0]==0:     #找到第一次找到关键词
            info["胸透"][0],info["胸透"][1]=1,j               
            index_list.append(j)
        elif bool([i for i in ECG_pattern  if i in sentence[j]]) and info["心电图"][0]==0:     #找到第一次找到关键词
            info["心电图"][0],info["心电图"][1]=1,j               
            index_list.append(j)
        elif bool([i for i in ALT_pattern  if i in sentence[j]]) and info["谷丙转氨酶"][0]==0:     #找到第一次找到关键词
            info["谷丙转氨酶"][0],info["谷丙转氨酶"][1]=1,j               
            index_list.append(j)
        elif bool([i for i in AST_pattern  if i in sentence[j]]) and info["谷草转氨酶"][0]==0:     #找到第一次找到关键词
            info["谷草转氨酶"][0],info["谷草转氨酶"][1]=1,j               
            index_list.append(j)
        elif bool([i for i in GGT_pattern  if i in sentence[j]]) and info["谷氨酰转移酶"][0]==0:     #找到第一次找到关键词
            info["谷氨酰转移酶"][0],info["谷氨酰转移酶"][1]=1,j               
            index_list.append(j)
        elif bool([i for i in internal_pattern  if i in sentence[j]]) and info["内科"][0]==0:     #找到第一次找到关键词
            info["内科"][0],info["内科"][1]=1,j               
            index_list.append(j)
        elif bool([i for i in surgery_pattern  if i in sentence[j]]) and info["外科"][0]==0:     #找到第一次找到关键词
            info["外科"][0],info["外科"][1]=1,j               
            index_list.append(j)
        elif bool([i for i in BP_pattern  if i in sentence[j]]) and info["血压"][0]==0:     #找到第一次找到关键词
            info["血压"][0],info["血压"][1]=1,j               
            index_list.append(j)
        elif bool([i for i in mouth_pattern  if i in sentence[j]]) and info["咽部"][0]==0:     #找到第一次找到关键词
            info["咽部"][0],info["咽部"][1]=1,j               
            index_list.append(j)
        elif bool([i for i in summary  if i in sentence[j]]) and info["医生总结"][0]==0:     #找到第一次找到关键词
            info["医生总结"][0],info["医生总结"][1]=1,j               
            index_list.append(j)    

    sorted_list = sorted(index_list)        #进行排序


    for key, value in info.items():
        if value[0]==0:
            print(f"{key}项目未被检测到")
        else :
            next_index = sorted_list.index(value[1])+1
            if next_index < len(sorted_list):value[2]=sorted_list[next_index]       
            else:value[2]=-1         #最后一个段落就是直接默认末尾位置为-1
            

    print(f"最终得到的字典为{info},排序后的结果为{sorted_list}")

#找指标对应的值
def searcher_value(sentences,traget):
    value,max_value=0,0
    for i in range(1,4):  #数值顶多后退三位
        print(f"数值--现在正在检验{sentences[sentences.index(traget)+i]}")
        if sentences[sentences.index(traget)+i][0].isdigit() and value==0:       #以数字开头的为数值的可能性更大
            value = eval(re.search(r'\d+(\.\d+)?', sentences[sentences.index(traget)+i])[0])        #判断是否含有整数或者小数,并得到ALT的值
            for j in range(1,4):  #范围在数值的下标下顶多后退三位
                print(f"范围--现在正在检验{sentences[sentences.index(traget)+i+j]}")
                if bool(re.search(r'\d+(\.\d+)?', sentences[sentences.index(traget)+i+j])):                      #存在数值就后续考虑
                    ranges = re.split('[{}]'.format(re.escape("-~--")), sentences[sentences.index(traget)+i+j])    #这里范围暂时只用~-作为分隔符
                    ranges = [i for i in ranges if len(i)>0]            #防止--连在一块
                    if len(ranges)>=2:          #有分隔符的情况
                        print(f"提取出的范围是{ranges},提取出的值是{value}")
                        print(ranges[1])
                        max_value = eval(ranges[1])        #判断是否含有整数或者小数
                        break
                    elif len(ranges)==1 and ranges[0].isdigit():       #分隔符没有识别出来
                        max_value = eval(ranges[0][-2:])        #判断是否含有整数或者小数
                        break
    
    return value,max_value


def pro_keyword(txt,key,key_words,direction):       #1表示正向，0表示反向
    if direction:           #正向，有的话说明是异常的
        for j in range(info[key][1],info[key][2]):
            worrys= [i for i in key_words  if i in txt[j]]      #获取异常情况
            if bool(worrys):
                print(f"{key}项目存在异常，异常为{worrys}")
                break
        if j==info[key][2]-1: print(f"{key}项目正常")
    else:                   #正向，有的话说明是异常的
        #先看项目里有没有结果没有就跳
        flag = 0;txt_val = 0
        for i in range(info[key][1],info[key][2]):  
            if (TXT_location[i][0][0] >= txt_val-5 and TXT_location[i][0][0] <= txt_val+5) and flag  :         #防止定位有波动，给5的容忍度
                print(f"获得结果：{txt[i]}---位置为{TXT_location[i][0][0]}")
                if txt[i] not in key_words:
                    print(f"{key}项目结果不满足正常情况，异常为{txt[i]}")
                    break
            if not flag and "结果" in txt[i]:          #找到结果值了,（防止有多个结果值，我只看第一个）
                flag = 1
                print(f"获得定位值{txt[i]}---位置为{TXT_location[i][0][0]}")
                txt_val = TXT_location[i][0][0]          #获得最左边的值
        if flag==0:
            print(f"{key}项目未能找到结果,建议看医生建议")



#检测是否有异常
def is_worry(sentences):
    ####数值比较类(但是这里都是默认指标下一位就是指标值，这个做法其实不一定对)
    #血压类
    SBP = sentences[sentences.index("收缩压")+1] 
    DBP = sentences[sentences.index("舒张压")+1] 
    if SBP.isdigit() and DBP.isdigit():
        print(f"SBP:{SBP},DBP:{DBP}")
        if 90<=eval(SBP)<=140 and 60<=eval(DBP)<=90:
            print("血压异常")
    else:
        print(f"读取错误！！SBP:{SBP},DBP:{DBP}")

    #肝功能类

    ALT,AST,GGT=0,0,0
    ALT_list,AST_list,GGT_list=0,0,0
    for s in sentences:             
        if ALT*AST*GGT!=0:break
        if bool([i for i in ALT_pattern  if i in s]) and ALT==0:         #找到文中的符合描述
            ALT,max_ALT=searcher_value(sentences,s)
            print(f"{s}---{ALT}---{max_ALT}")
        if bool([i for i in AST_pattern  if i in s]) and AST==0:
            AST ,max_AST=searcher_value(sentences,s)
            print(f"{s}---{AST}---{max_AST}")
        if bool([i for i in GGT_pattern  if i in s]) and GGT==0: 
            GGT ,max_GGT=searcher_value(sentences,s)
            print(f"{s}---{GGT}---{max_GGT}")


    ####文字抓取类
    #正向【出现关键字，状态码设置成0】
    eye = ["视力高度近视","散光","弱视","色弱","色盲","红绿色弱","红绿色盲"]
    if info["视力"][0]!=0: pro_keyword(sentences,"视力",eye,1)

    

    #反向【出现关键字，状态码设置成1】
    inside = ["心肺无异常","肝脾无肿大","腹部无包块","无异常","未见明显异常","正常","未见异常","无明显异常"]
    if info["内科"][0]!=0: pro_keyword(sentences,"内科",inside,0)

    outside = ["淋巴结无肿大","甲状腺无异常","乳房无异常","外生殖器无异常","前列腺无异常","四肢脊柱无畸形","活动正常","正常","未见异常","无异常","未见明显异常","无明显异常"]
    if info["外科"][0]!=0: pro_keyword(sentences,"外科",outside,0)

    ear_nose_mouth = ["正常","未见异常","无异常","未见明显异常","无明显异常"]
    if info["听力"][0]!=0: pro_keyword(sentences,"听力",ear_nose_mouth,0)
    if info["鼻"][0]!=0: pro_keyword(sentences,"鼻",ear_nose_mouth,0)
    if info["咽部"][0]!=0: pro_keyword(sentences,"咽部",ear_nose_mouth,0)
    if info["胸透"][0]!=0: pro_keyword(sentences,"胸透",ear_nose_mouth,0)

    ECG = ["无异常","窦性心律","窦性心律不齐","心律不齐","正常心电图","偏正常心电图","正常","未见异常","未见明显异常","无明显异常"]
    if info["心电图"][0]!=0: pro_keyword(sentences,"心电图",ECG,0)

    conclusion = ["正常","未见异常","无异常","未见明显异常","无明显异常","符合入职标准","不影响正常工作"]
    if info["医生总结"][0]!=0: pro_keyword(sentences,"医生总结",conclusion,0)

#主程序
def main(imgs_path):
    print(f"现在主程序里传入的是{imgs_path}")
    result,content,location,flag = img2word(imgs_path)
    if flag > 0.1:
        print(f"##############################该简历质量不高,建议转人工处理！识别率仅为{1-flag}###################################")
    print(result)
    global TXT_location 
    TXT_location  = location
    is_enough(content)
    is_worry(content)
    

if __name__=="__main__":
    # pdf_path = r"C:\Users\Administrator\Desktop\体检报告\二期\10026937-TJ-02.jpg"
    pdf_path = r"C:\Users\Administrator\Desktop\体检\体检\10026800-TJ-01xy.pdf"
    if pdf_path.endswith(".pdf"):           
        imgs_path_list = extract_text_allpage(pdf_path,"img")            #得到是简历解析成图像的路径，所以不用担心会多读
        main(imgs_path_list)
    elif pdf_path.endswith(".jpg") or pdf_path.endswith(".png"):
        imgs_path = pdf_path
        main(imgs_path)
    else:
        print("不支持的文件类型！")


    # 读取文件
    # # 加载工作簿
    # workbook = load_workbook(filename=r'C:\Users\Administrator\Desktop\预警.xlsx')
    # # 选择工作表
    # sheet = workbook.active
    # # 遍历每个单元格
    # info = dict()
    # for row in sheet.iter_rows(values_only=False):
    #     index = 0
    #     cur_key = None
    #     for cell in row:
    #         if type(cell.value) is not None and index == 0 :
    #             cur_key = cell.value
    #             info[cur_key]=[]
    #         elif index == 1 and cur_key is not None:            #加入体检项目关键词【后续还要加入体检项目异常值】
    #             info[cur_key].append(cell.value)
    #         index+=1
    # print(info)




